# pylint: disable = invalid-name, superfluous-parens, bad-whitespace, using-constant-test
"""
Take a multisheet xlsx and for each sheet break its compound statements into more rows of data
Usage: python break_compounds.py
"""
import sys
import copy
import nltk
#nltk.download('punkt')
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
#import coding_helper

input_file = sys.argv[1]
output_file = sys.argv[2]

### HELPERS

def get_col_headers(ws):
    """
    for excel style worksheet, get dict mapping col names to numeric indices for use in openpyxl
    """
    header_row = ws[1]
    # extract value of first row (header string name)
    row_headers = [cell.value for cell in header_row]
    # extract column name (excel standard alpha code)
    row_cols = [cell.column for cell in header_row]
    # turn alpha code to number of col
    row_idxs = [column_index_from_string(cell)-1 for cell in row_cols]
    # keys are familiar header names, values are numeric indices, 0-indexed
    ref = dict(zip(row_headers, row_idxs))
    return(ref)

def get_col_headers_workbook(workbook):
    """
    for workbook whose sheets all have the same columns,
    get dict mapping col names to numeric indices for use in openpyxl
    """
    ws1 = workbook.active
    headers = get_col_headers(ws1)
    return( headers )

if __name__ == "__main__":
    wb = load_workbook(input_file)
    headerRef = get_col_headers_workbook(wb)
    #print( headerRef )

    rows_to_change = []
    for sheet in wb.worksheets:
        if not sheet.title.startswith( 'statements_' ):
            continue
        print( 'SHEET', sheet.title )
        rows_original = 0
        rows_added = 0
        #print(sheet)
        for i, row in enumerate(sheet.iter_rows(), start=1):
            #assert( row == sheet[i],
            #   "enumerate here is supposed to give the index of the row in sheet" )
            an_is = row[ headerRef['Institutional Statement'] ].value
            if an_is is None:
                break # stop at the end of the sheet
            an_is = an_is.replace(r'\\', ' ') # for minecraft
            if an_is.startswith('='):
                an_is = '\\' + an_is
                an_is = an_is.replace(r'.**', '. ') # for reddit: sentences followed by markdown
            an_is = an_is.replace(r'.*', '. ') # for reddit
            an_is = an_is.replace(r'\n\n**', '. ') # for reddit: double \n+bullets instead of .'s
            an_is = an_is.replace(r'\n\n*', '. ') # for reddit
            an_is = an_is.replace(r'\n\n-', '. ') # for reddit
            an_is = an_is.replace(r'\n**', '. ') # for reddit
            an_is = an_is.replace(r'\n*', '. ') # for reddit
            an_is = an_is.replace(r'\n', ' ') # for reddit
            iss = nltk.sent_tokenize( an_is )
            if len( iss ) == 1: # non-compound statement
                row[ headerRef['Institutional Statement'] ].value = an_is
            else: # compound statement
                #print()
                # rows_to_change is the rows to be cahnged in the orignal sheet
                #   old compounds should be removed, while new ones should be added
                #   each element is a 3-tuple: row #, add/remove, row object
                #   The order that things get added to rows to change matters:
                #   The old row-to-be-deleted has to get appended after its parts-to-be-added
                for an_is_part in iss:
                    #print(i, an_is_part )
                    # This copy is important.
                    # Both the row and cells have to be deep copied
                    new_row = [copy.copy(cell) for cell in row]
                    new_row[ headerRef['Institutional Statement'] ].value = an_is_part
                    new_row[ headerRef['Text Type'] ].value = 'decompounded'
                    rows_to_change.append( (i, 1, new_row) )
                    rows_added += 1
                rows_to_change.append( (i, 0, row) )
                rows_added -= 1
        rows_original = i  #pylint: disable=undefined-loop-variable
        # this reverse is important.
        #   it makes the row numbers stay the same for each row that has to be changed.
        rows_to_change.reverse()
        # perform changes
        for i, action, row in rows_to_change:
            #print(i, action, row)
            if action:
                sheet.insert_rows(i)
                # have to go cell by cell
                #  from https://stackoverflow.com/questions/33920108/writing-to-row-using-openpyxl
                for col, cell in enumerate(row, start=1):
                    sheet.cell(row=i, column=col).value = cell.value
                    sheet.cell(row=i, column=col).style = cell.style
            else:
                sheet.delete_rows(i)
        # Print it out
        print("Rows starting: {} Rows added: {}".format( rows_original, rows_added ))
        if False:
            for i, row in enumerate(sheet.iter_rows(), start=1):
                an_is = row[ headerRef['Institutional Statement'] ].value
                if an_is is None:
                    break # stop at the end of the sheet
                print( i, an_is )
        ### now do sheet level edits
        # extend data validations to new rows
        for validation in sheet.data_validations.dataValidation:
            validation.sqref.ranges[0].expand(down=rows_added)
        ### currently only works for one sheet.  some bug makes it fail for others.
        break
    wb.save( output_file )
